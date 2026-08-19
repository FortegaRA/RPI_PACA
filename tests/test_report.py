"""
Tests for the xlsx executive-summary report.

The pipeline is local-only: Notion/Airtable/SMTP delivery were removed on
16/08/2026 (company compliance forbids automated connectors), so the report file
itself is the hand-off. Nothing here touches the network.
"""

import csv
import os
import sys
import tempfile
import unittest

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)
sys.path.insert(0, os.path.join(ROOT, "tests"))

import consolidate  # noqa: E402
import report  # noqa: E402


def row(rs, cc, term, rtype="APPROVAL", appr=None, name="P", applicant="A"):
    base = {f: None for f in report.CANONICAL_FIELDS}
    base.update({"registration_number": rs, "country_code": cc, "record_type": rtype,
                 "molecule_search_term": term, "approval_date": appr,
                 "product_name": name, "applicant": applicant, "extraction_date": "14/06/2026"})
    return base


SAMPLE = [
    row("R1", "US", "REGORAFENIB", appr="01/06/2026", name="STIVARGA"),
    row("R2", "SV", "DAPAGLIFLOZINA", appr="10/03/2020", name="FORXIGA"),
    row("R3", "CO", "EMPAGLIFLOZINA", rtype="SUBMISSION"),
    row("NO-DATA-SV-MARSTACIMAB", "SV", "MARSTACIMAB", rtype="NO_DATA"),
]


# ── Report ────────────────────────────────────────────────────────────────────
class TestReport(unittest.TestCase):
    def test_build_report_has_expected_sheets(self):
        from openpyxl import load_workbook
        with tempfile.TemporaryDirectory() as t:
            path = report.generate(t, rows=SAMPLE)
            self.assertTrue(path and os.path.exists(path))
            wb = load_workbook(path)
            self.assertIn("Resumen", wb.sheetnames)
            self.assertIn("Datos", wb.sheetnames)
            # Datos has a header row + one row per record (incl. NO_DATA).
            self.assertEqual(wb["Datos"].max_row, len(SAMPLE) + 1)

    def test_generate_none_on_empty(self):
        with tempfile.TemporaryDirectory() as t:
            self.assertIsNone(report.generate(t, rows=[]))

    def test_recent_approvals_filters_by_date(self):
        recents = report.recent_approvals(SAMPLE, days=90)
        names = {r["product_name"] for r in recents}
        self.assertIn("STIVARGA", names)       # 01/06/2026 — recent
        self.assertNotIn("FORXIGA", names)      # 10/03/2020 — old

    def test_text_summary_counts_exclude_no_data(self):
        text = report.text_summary(SAMPLE)
        self.assertIn("2 approvals", text)      # NO_DATA not counted as approval
        self.assertIn("1 submissions", text)
        self.assertIn("MARSTACIMAB", text)      # listed as "no registrations found"
        self.assertIn("STIVARGA", text)         # recent approval surfaced


class TestPicksNewestConsolidated(unittest.TestCase):
    """The stamp is DDMMYYYY, so alphabetical order is NOT chronological order.

    Regression: "28072026" (28 Jul) sorts after "16082026" (16 Aug) as plain text,
    so the report used to build itself from a stale consolidated file.
    """

    def _make(self, folder, stamp, marker):
        path = os.path.join(folder, f"RPI_CONSOLIDATED_{stamp}.csv")
        with open(path, "w", encoding="utf-8-sig", newline="") as fh:
            fh.write("registration_number,country_code,record_type,product_name\n")
            fh.write(f"R1,CR,APPROVAL,{marker}\n")
        return path

    def test_day_first_stamp_sorts_chronologically(self):
        with tempfile.TemporaryDirectory() as t:
            self._make(t, "28072026", "JULIO")     # alphabetically last
            newest = self._make(t, "16082026", "AGOSTO")   # chronologically newest
            self.assertEqual(report.latest_consolidated(t), newest)
            rows = report.load_consolidated(t)
            self.assertEqual(rows[0]["product_name"], "AGOSTO")

    def test_single_file_is_returned(self):
        with tempfile.TemporaryDirectory() as t:
            only = self._make(t, "13062026", "UNICO")
            self.assertEqual(report.latest_consolidated(t), only)

    def test_no_files_returns_none(self):
        with tempfile.TemporaryDirectory() as t:
            self.assertIsNone(report.latest_consolidated(t))
            self.assertEqual(report.load_consolidated(t), [])

    def test_unparseable_stamp_never_beats_a_dated_file(self):
        with tempfile.TemporaryDirectory() as t:
            self._make(t, "BORRADOR", "BASURA")
            good = self._make(t, "13062026", "BUENO")
            self.assertEqual(report.latest_consolidated(t), good)



class TestReportCoversMergedData(unittest.TestCase):
    """The report must describe the same dataset the consolidated CSV does.

    Regression: the report used to be built from the rows extracted in the current
    run. When a source came back empty, merge-don't-clobber kept its previous CSV
    on disk (so it stayed in the consolidated) but it contributed 0 in-memory rows —
    and the whole country vanished from the workbook with no warning. A real run on
    18/08/2026 produced a report with 8 countries and no Colombia while the
    consolidated correctly carried its 656 rows.
    """

    def test_generate_without_rows_reads_the_consolidated(self):
        with tempfile.TemporaryDirectory() as t:
            path = os.path.join(t, "RPI_CONSOLIDATED_18082026.csv")
            with open(path, "w", encoding="utf-8-sig", newline="") as fh:
                w = csv.DictWriter(fh, fieldnames=report.CANONICAL_FIELDS,
                                   extrasaction="ignore")
                w.writeheader()
                w.writerow(row("R-CO", "CO", "DAPAGLIFLOZINA", name="FORXIGA"))
                w.writerow(row("R-US", "US", "REGORAFENIB", name="STIVARGA"))
            out = report.generate(t)                      # no rows= -> consolidated
            self.assertTrue(out and os.path.exists(out))
            from openpyxl import load_workbook
            wb = load_workbook(out)
            countries = {wb["Matriz Molécula×País"].cell(1, c).value
                         for c in range(2, wb["Matriz Molécula×País"].max_column + 1)}
            self.assertIn("CO", countries, "a merged-only country must still be reported")
            self.assertIn("US", countries)



class TestStableConsolidatedName(unittest.TestCase):
    """The consolidated file keeps ONE name so a fixed SharePoint/Fabric path works.

    A dated filename forced whoever uploads it to rename the file before every
    upload — a manual step that eventually gets done wrong and silently breaks the
    downstream refresh.
    """

    def _write(self, path, marker):
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "w", encoding="utf-8-sig", newline="") as fh:
            w = csv.DictWriter(fh, fieldnames=report.CANONICAL_FIELDS,
                               extrasaction="ignore")
            w.writeheader()
            w.writerow(row("R1", "CR", "REGORAFENIB", name=marker))

    def test_stable_file_wins_over_dated_leftovers(self):
        with tempfile.TemporaryDirectory() as t:
            self._write(os.path.join(t, "RPI_CONSOLIDATED_28072026.csv"), "VIEJO")
            stable = os.path.join(t, consolidate.CONSOLIDATED_NAME)
            self._write(stable, "ACTUAL")
            self.assertEqual(report.latest_consolidated(t), stable)
            self.assertEqual(report.load_consolidated(t)[0]["product_name"], "ACTUAL")

    def test_dated_files_still_readable_before_the_rename(self):
        """Output folders written by an older version must keep working."""
        with tempfile.TemporaryDirectory() as t:
            old = os.path.join(t, "RPI_CONSOLIDATED_28072026.csv")
            self._write(old, "VIEJO")
            self.assertEqual(report.latest_consolidated(t), old)

    def test_archived_history_is_found_as_a_fallback(self):
        with tempfile.TemporaryDirectory() as t:
            arch = os.path.join(t, consolidate.HISTORY_DIRNAME,
                                "RPI_CONSOLIDATED_13062026.csv")
            self._write(arch, "ARCHIVADO")
            self.assertEqual(report.latest_consolidated(t), arch)


class TestConsolidateWritesStableAndHistory(unittest.TestCase):
    def test_writes_stable_name_plus_dated_archive(self):
        from datetime import date as _date
        with tempfile.TemporaryDirectory() as t:
            for cc in ("CR", "HN"):
                p = os.path.join(t, f"RPI{cc}_canonical.csv")
                with open(p, "w", encoding="utf-8-sig", newline="") as fh:
                    w = csv.DictWriter(fh, fieldnames=report.CANONICAL_FIELDS,
                                       extrasaction="ignore")
                    w.writeheader()
                    w.writerow(row(f"R-{cc}", cc, "REGORAFENIB"))
            out = consolidate.consolidate(t, quiet=True)
            self.assertEqual(os.path.basename(out), consolidate.CONSOLIDATED_NAME)
            self.assertTrue(os.path.exists(out))
            dated = os.path.join(t, consolidate.HISTORY_DIRNAME,
                                 f"RPI_CONSOLIDATED_{_date.today():%d%m%Y}.csv")
            self.assertTrue(os.path.exists(dated), "una copia fechada debe archivarse")

    def test_consolidated_is_never_re_ingested(self):
        """Running twice must not double the rows."""
        with tempfile.TemporaryDirectory() as t:
            for cc in ("CR", "HN"):
                p = os.path.join(t, f"RPI{cc}_canonical.csv")
                with open(p, "w", encoding="utf-8-sig", newline="") as fh:
                    w = csv.DictWriter(fh, fieldnames=report.CANONICAL_FIELDS,
                                       extrasaction="ignore")
                    w.writeheader()
                    w.writerow(row(f"R-{cc}", cc, "REGORAFENIB"))
            consolidate.consolidate(t, quiet=True)
            out = consolidate.consolidate(t, quiet=True)
            with open(out, encoding="utf-8-sig") as fh:
                self.assertEqual(len(list(csv.DictReader(fh))), 2)


if __name__ == "__main__":
    unittest.main(verbosity=2)
