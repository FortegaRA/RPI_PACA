"""
hn.py — Honduras ARSA bulk Excel download + openpyxl parse.

Portal: https://sicreb.arsa.hn/Servicios/Medicamentos?...
ARSA publishes its full active-registration list as a single Excel export. The
portal page (confirmed live) links it directly as ``DescargarExcel?NombrePestania=MRS``,
so we download it over plain HTTP first and only fall back to the Selenium
click-the-button flow if the direct download fails. The workbook's "Registros"
sheet is parsed with openpyxl, keeping rows whose active-substance column
contains any target molecule term.

record_type and status are hardcoded (APPROVAL / APPROVED) — ARSA only publishes
active registrations.
"""

from __future__ import annotations

import os
import re
import time
from urllib.parse import urljoin

import config
from extractors import selenium_base as sb

COUNTRY_CODE = "HN"
SOURCE_URL = config.SOURCE_URLS["hn"]
EXPECTED_FILE = "MedicamentosArsa.xlsx"
DIRECT_DOWNLOAD_URL = "https://sicreb.arsa.hn/Servicios/DescargarExcel?NombrePestania=MRS"

# Header name -> candidate spellings in the Excel sheet.
_COLS = {
    "registration_number": ["Numero de Registro Sanitario", "Número de Registro Sanitario",
                            "No. Registro Sanitario", "Registro Sanitario"],
    "product_name": ["Nombre del producto", "Nombre del Producto", "Producto"],
    "api": ["Nombre de sustancias activas / Principio activo", "Principio activo",
            "Nombre de sustancias activas", "Sustancias activas"],
    "applicant": ["Nombre del titular", "Titular"],
    "manufacturer": ["Nombre del fabricante", "Fabricante"],
    "manufacturer_country": ["Pais del Fabricante", "País del Fabricante", "Pais Fabricante"],
    "dosage_form": ["Forma farmacéutica", "Forma farmaceutica", "Forma Farmacéutica"],
    "concentration": ["Concentracion por unidad de dosis", "Concentración por unidad de dosis",
                      "Concentracion", "Concentración"],
    "process_type": ["Tipo de solicitud", "Tipo de Solicitud"],
    # Not canonical — used only by post-filters (e.g. tadalafil PAH vs ED).
    "_indication_hint": ["Grupo terapéutico", "Grupo terapeutico", "Grupo Terapéutico"],
}

# ARSA's Excel carries no approval-date column, but the registration number encodes
# it: HN-M-MMYY-NNNN, e.g. HN-M-0723-0046 -> July 2023. Verified across all 72
# HN-prefixed rows: the first pair spans 01-12 (all twelve months present) and the
# second only 18-25 — years, not days. There is therefore NO day component, so the
# derived date is anchored to the 1st and is month-precision. Anchoring at the start
# of the month never makes a registration look more recent than it is.
#
# 30 further rows use a legacy 5-digit number (45992) that encodes nothing; those keep
# an empty approval_date rather than an invented one.
_REG_DATE = re.compile(r"^HN-[A-Z]+-(\d{2})(\d{2})-", re.I)
_REG_DATE_CENTURY = 2000


def approval_date_from_registration(registration) -> str | None:
    """Derive DD/MM/YYYY (day anchored to 01) from an ARSA registration number.

    Returns ``None`` when the number carries no date — an out-of-range month, the
    legacy 5-digit format, or anything unrecognized.
    """
    m = _REG_DATE.match(str(registration or "").strip())
    if not m:
        return None
    month, year = int(m.group(1)), int(m.group(2))
    if not 1 <= month <= 12:
        return None
    return f"01/{month:02d}/{_REG_DATE_CENTURY + year}"


_DOWNLOAD_XPATHS = [
    "//a[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'excel')]",
    "//button[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'excel')]",
    "//a[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'descargar')]",
    "//button[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'exportar')]",
    "//*[@title and contains(translate(@title, 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'excel')]",
    "//i[contains(@class,'fa-file-excel')]/ancestor::a[1]",
]


def _norm_header(text) -> str:
    """Normalize a header for matching: lowercase, collapse spaces, drop trailing ':'."""
    return re.sub(r"\s+", " ", str(text).strip().lower()).rstrip(":").strip()


def _resolve_headers(header_row: list) -> dict:
    # ARSA headers carry trailing colons ("Nombre del producto:"), so match on
    # normalized text with exact-or-substring rather than an exact dict lookup.
    headers = [(_norm_header(h), idx) for idx, h in enumerate(header_row)
               if h not in (None, "")]
    resolved: dict = {}
    for field, candidates in _COLS.items():
        for cand in candidates:
            cn = _norm_header(cand)
            hit = next((idx for htext, idx in headers if htext == cn or cn in htext), None)
            if hit is not None:
                resolved[field] = hit
                break
    return resolved


def _parse_workbook(path: str, molecules: list[dict]) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = wb["Registros"] if "Registros" in wb.sheetnames else wb[wb.sheetnames[0]]

    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header = list(next(rows_iter))
    except StopIteration:
        return []
    resolved = _resolve_headers(header)
    api_idx = resolved.get("api")

    # latam_term + INN-variant aliases (broad class terms like "FACTOR VIII" excluded).
    terms = [(m, [t.lower() for t in config.search_terms(m)]) for m in molecules]

    out = []
    for raw in rows_iter:
        if api_idx is None or api_idx >= len(raw):
            continue
        api_val = str(raw[api_idx] or "").lower()
        matched = next((m for m, ts in terms if any(t in api_val for t in ts)), None)
        if matched is None:
            continue
        row = {field: (raw[idx] if idx < len(raw) else None)
               for field, idx in resolved.items()}
        row.update({
            "country_code": COUNTRY_CODE,
            "record_type": "APPROVAL",
            "status": "APPROVED",
            "molecule_search_term": matched["latam_term"].upper(),
            "source_url": SOURCE_URL,
        })
        # ARSA publishes no approval-date column; the registration number encodes
        # month/year. Only fill what the sheet left blank.
        if not row.get("approval_date"):
            derived = approval_date_from_registration(row.get("registration_number"))
            if derived:
                row["approval_date"] = derived
        out.append(row)
    wb.close()
    return out


def _download_direct(download_dir: str, errors: list) -> str | None:
    """Fetch the ARSA Excel over plain HTTP (2 attempts). Returns path or None."""
    target = os.path.join(download_dir, EXPECTED_FILE)
    session = config.build_session()
    for attempt in (1, 2):
        try:
            # Visit the page first (cookies) and pick up the real link if it moved.
            url = DIRECT_DOWNLOAD_URL
            try:
                page = session.get(SOURCE_URL, timeout=30)
                match = re.search(r'href="([^"]*DescargarExcel[^"]*)"', page.text)
                if match:
                    url = urljoin(SOURCE_URL, match.group(1).replace("&amp;", "&"))
            except Exception:
                pass
            with session.get(url, timeout=180, stream=True) as resp:
                resp.raise_for_status()
                ctype = resp.headers.get("Content-Type", "").lower()
                if "html" in ctype:
                    raise ValueError(f"expected an Excel payload, got {ctype}")
                with open(target, "wb") as fh:
                    for chunk in resp.iter_content(chunk_size=1 << 16):
                        fh.write(chunk)
            if os.path.getsize(target) > 10_000:  # sanity: not an error stub
                return target
            raise ValueError("downloaded file is suspiciously small")
        except Exception as exc:  # noqa: BLE001
            msg = sb.err_line(exc)
            print(f"    [HN] direct download attempt {attempt} failed: {msg}")
            if attempt == 2:
                errors.append(f"HN: direct download failed: {msg}")
            time.sleep(5)
    return None


def _download_selenium(download_dir: str, headless: bool, errors: list) -> str | None:
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait

    driver = None
    try:
        driver = sb.build_driver(headless=headless, download_dir=download_dir)
        driver.get(SOURCE_URL)
        WebDriverWait(driver, 30).until(lambda d: d.execute_script(
            "return document.readyState") == "complete")

        clicked = False
        for xpath in _DOWNLOAD_XPATHS:
            try:
                btn = driver.find_element(By.XPATH, xpath)
                driver.execute_script("arguments[0].click();", btn)
                clicked = True
                break
            except Exception:
                continue
        if not clicked:
            print("    [HN] could not locate the Excel download control")
            errors.append("HN: Excel download control not found")
            return None

        path = sb.wait_for_download(download_dir, EXPECTED_FILE, timeout=120,
                                    exts=(".xlsx", ".xls"))
        if not path:
            # Fall back to newest xlsx in the folder.
            path = sb.wait_for_download(download_dir, None, timeout=10, exts=(".xlsx", ".xls"))
        if not path or not os.path.exists(path):
            print("    [HN] download did not complete")
            errors.append("HN: Selenium download did not complete")
            return None
        return path
    except Exception as exc:  # noqa: BLE001
        msg = sb.err_line(exc)
        print(f"    [HN] Selenium download failed: {msg}")
        errors.append(f"HN: {msg}")
        return None
    finally:
        if driver is not None:
            driver.quit()


def extract(molecules: list[dict], config_dict: dict) -> list[dict]:
    errors = config_dict.setdefault("partial_errors", [])
    download_dir = config_dict.get("download_dir") or config_dict.get("output_dir")
    headless = config_dict.get("headless", False)
    sb.clear_downloads(download_dir, EXPECTED_FILE)

    path = _download_direct(download_dir, errors)
    if path is None and not config_dict.get("no_selenium"):
        print("    [HN] falling back to Selenium download")
        path = _download_selenium(download_dir, headless, errors)
    if path is None:
        return []

    rows = _parse_workbook(path, molecules)
    print(f"    [HN] {len(rows)} matching registrations from {os.path.basename(path)}")
    return rows
