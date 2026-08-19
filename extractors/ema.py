"""
ema.py — European Medicines Agency (reference / forecast source).

Method: plain HTTP GET of the EMA "Medicine" output **XLSX**, filtered in memory.
No browser. (EMA retired the old per-report JSON feeds in its 2026 site redesign;
both `…_json-report_en.json` URLs now 404. The replacement is a single spreadsheet
that lists Human + Veterinary medicines with a `Medicine status` column.)

We keep only Human rows, match a molecule when its INN / active substance contains
the molecule's EMA term (or an alias), and split APPROVAL vs SUBMISSION by whether
the medicine ever received a marketing authorisation:

    has a Marketing authorisation / EC decision date  -> APPROVAL
    otherwise (Opinion, Refused, Application withdrawn) -> SUBMISSION
"""

from __future__ import annotations

import io

import config

COUNTRY_CODE = "EU"

# Header labels in the XLSX (normalized: lowercased, whitespace-collapsed). The
# sheet carries newlines inside some labels, so we always match on the normalized
# form rather than the raw string.
_H_CATEGORY = "category"
_H_NAME = "name of medicine"
_H_NUMBER = "ema product number"
_H_STATUS = "medicine status"
_H_INN = "international non-proprietary name (inn) / common name"
_H_SUBSTANCE = "active substance"
_H_AREA = "therapeutic area (mesh)"
_H_INDICATION = "therapeutic indication"
_H_BIOSIMILAR = "biosimilar"
_H_APPLICANT = "marketing authorisation developer / applicant / holder"
_H_EC_DATE = "european commission decision date"
_H_OPINION_DATE = "opinion adopted date"
_H_EVAL_DATE = "start of evaluation date"
_H_MA_DATE = "marketing authorisation date"
_H_EXPIRY_DATE = "withdrawal / expiry / revocation / lapse of marketing authorisation date"
_H_URL = "medicine url"


def _norm(text) -> str:
    """Normalize a header/category cell: lowercase, collapse all whitespace."""
    return " ".join(str(text or "").split()).lower()


def _build_terms(molecules: list[dict]) -> list[tuple]:
    """Return (molecule, [lowercased terms]) for matching on INN / substance."""
    out = []
    for m in molecules:
        terms = [m["ema_term"]] + list(m.get("aliases", []))
        out.append((m, [t.lower() for t in terms if t]))
    return out


def _match_molecule(text: str, term_table: list[tuple]):
    low = text.lower()
    for m, terms in term_table:
        if any(t in low for t in terms):
            return m
    return None


def _fetch_records(session, url: str, errors: list) -> list[dict] | None:
    """Download the XLSX and return a list of dict rows keyed by normalized header.

    Returns ``None`` on a fetch/parse failure (caller records the error), or a list
    (possibly empty) on success.
    """
    try:
        resp = session.get(url, timeout=60)
        resp.raise_for_status()
        content = resp.content
    except Exception as exc:  # noqa: BLE001 - report and continue
        print(f"    [EMA] fetch failed for {url.rsplit('/', 1)[-1]}: {exc}")
        errors.append(f"EMA: fetch failed for {url.rsplit('/', 1)[-1]}: {exc}")
        return None

    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - openpyxl ships with the project
        print(f"    [EMA] openpyxl unavailable: {exc}")
        errors.append(f"EMA: openpyxl unavailable: {exc}")
        return None

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        grid = [list(r) for r in ws.iter_rows(values_only=True)]
    except Exception as exc:  # noqa: BLE001
        print(f"    [EMA] could not parse XLSX: {exc}")
        errors.append(f"EMA: could not parse XLSX: {exc}")
        return None

    # Find the header row: the first row that carries both the product-number and
    # name columns (a banner row precedes it, and EMA may add/remove banner lines).
    header_idx = None
    for i, row in enumerate(grid):
        norm = {_norm(c) for c in row if c not in (None, "")}
        if _H_NUMBER in norm and _H_NAME in norm:
            header_idx = i
            break
    if header_idx is None:
        print("    [EMA] header row not found in XLSX — layout may have changed")
        errors.append("EMA: header row not found in XLSX")
        return None

    headers = [_norm(c) for c in grid[header_idx]]
    records: list[dict] = []
    for row in grid[header_idx + 1:]:
        if not any(c not in (None, "") for c in row):
            continue
        rec = {}
        for h, val in zip(headers, row):
            if h and h not in rec:
                rec[h] = val
        records.append(rec)
    return records


def _record_type_and_dates(rec: dict):
    """Decide APPROVAL vs SUBMISSION and pick the relevant dates."""
    ma_date = rec.get(_H_MA_DATE) or rec.get(_H_EC_DATE)
    if ma_date:
        return "APPROVAL", None, ma_date
    submission_date = rec.get(_H_OPINION_DATE) or rec.get(_H_EVAL_DATE)
    return "SUBMISSION", submission_date, None


def _indication_hint(rec: dict) -> str | None:
    parts = [rec.get(_H_INDICATION), rec.get(_H_AREA)]
    # The Biosimilar flag is decisive for the ranibizumab post-filter.
    if _norm(rec.get(_H_BIOSIMILAR)) == "yes":
        parts.append("biosimilar")
    text = " ".join(str(p) for p in parts if p)
    return text or None


def _harvest(records: list[dict], term_table: list[tuple]) -> list[dict]:
    rows = []
    for rec in records:
        if _norm(rec.get(_H_CATEGORY)) != "human":
            continue  # skip Veterinary
        inn = rec.get(_H_INN) or ""
        substance = rec.get(_H_SUBSTANCE) or ""
        molecule = _match_molecule(f"{inn} {substance}", term_table)
        if molecule is None:
            continue
        record_type, submission_date, approval_date = _record_type_and_dates(rec)
        rows.append({
            "registration_number": rec.get(_H_NUMBER),
            "product_name": rec.get(_H_NAME),
            "api": inn or substance,
            "applicant": rec.get(_H_APPLICANT),
            "status": rec.get(_H_STATUS),
            "approval_date": approval_date,
            "submission_date": submission_date,
            "expiration_date": rec.get(_H_EXPIRY_DATE),
            "source_url": rec.get(_H_URL) or config.SOURCE_URLS["ema"],
            "country_code": COUNTRY_CODE,
            "record_type": record_type,
            "molecule_search_term": molecule["ema_term"].upper(),
            "_indication_hint": _indication_hint(rec),
        })
    return rows


def extract(molecules: list[dict], config_dict: dict) -> list[dict]:
    errors = config_dict.setdefault("partial_errors", [])
    session = config_dict.get("session") or config.build_session()
    term_table = _build_terms(molecules)

    records = _fetch_records(session, config.SOURCE_URLS["ema_report"], errors)
    if records is None:
        return []

    human = sum(1 for r in records if _norm(r.get(_H_CATEGORY)) == "human")
    print(f"    [EMA] medicines report: {len(records)} records ({human} human)")
    rows = _harvest(records, term_table)
    print(f"    [EMA] matched {len(rows)} rows for target molecules")
    return rows
