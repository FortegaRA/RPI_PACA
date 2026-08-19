"""
report.py — Executive-summary xlsx report for the Regulatory Affairs team.

Turns the consolidated canonical rows into a competitor-intelligence workbook:

    Resumen                 KPIs, by-country / by-status breakdown, and the key
                            signal — competitor approvals in the last 90 days.
    Matriz Molécula×País    molecule × country activity matrix.
    Datos                   the full canonical table (autofilter, frozen header).

``text_summary`` produces the same headline numbers as plain text for the email
body. ``NO_DATA`` presence rows are reported separately and never counted as
approvals/submissions. openpyxl is imported lazily so this module stays import-safe.
"""

from __future__ import annotations

import csv
import glob
import os
from collections import Counter, defaultdict
from datetime import date, datetime

import consolidate
from normalize import CANONICAL_FIELDS

NO_DATA = "NO_DATA"
RECENT_DAYS = 90


# ── Data loading / small helpers ──────────────────────────────────────────────
def _consolidated_sort_key(path: str):
    """Sort key that orders RPI_CONSOLIDATED_DDMMYYYY.csv chronologically.

    The stamp is day-first, so a plain alphabetical sort is wrong: "28072026"
    (28 Jul) sorts after "16082026" (16 Aug) as text, which made the report build
    itself from a stale consolidated file. Files whose stamp cannot be parsed fall
    back to their modification time so they never win over a dated file by accident.
    """
    stem = os.path.basename(path).removeprefix("RPI_CONSOLIDATED_").removesuffix(".csv")
    try:
        return (1, datetime.strptime(stem, "%d%m%Y").date())
    except ValueError:
        try:
            return (0, date.fromtimestamp(os.path.getmtime(path)))
        except OSError:
            return (0, date.min)


def latest_consolidated(output_dir: str) -> str | None:
    """Path of the current consolidated file, or None.

    The live file has a stable name (``RPI_CONSOLIDATED.csv``) and always wins. The
    dated fallback covers output folders written before the rename, and the archived
    copies under ``_history/``.
    """
    stable = os.path.join(output_dir, consolidate.CONSOLIDATED_NAME)
    if os.path.exists(stable):
        return stable
    files = glob.glob(os.path.join(output_dir, "RPI_CONSOLIDATED_*.csv"))
    files += glob.glob(os.path.join(output_dir, consolidate.HISTORY_DIRNAME,
                                    "RPI_CONSOLIDATED_*.csv"))
    return max(files, key=_consolidated_sort_key) if files else None


def load_consolidated(output_dir: str) -> list[dict]:
    path = latest_consolidated(output_dir)
    if not path:
        return []
    with open(path, encoding="utf-8-sig", newline="") as fh:
        return [{k: (v if v not in ("", None) else None) for k, v in row.items()}
                for row in csv.DictReader(fh)]


def _parse(d):
    if not d:
        return None
    try:
        return datetime.strptime(str(d).strip(), "%d/%m/%Y").date()
    except ValueError:
        return None


def _split(rows):
    """Return (real_rows, presence_rows)."""
    real = [r for r in rows if (r.get("record_type") or "") != NO_DATA]
    presence = [r for r in rows if (r.get("record_type") or "") == NO_DATA]
    return real, presence


def _has_placeholder_date(row: dict) -> bool:
    """True when a row's approval_date is a stand-in stamped at extraction time.

    A row whose approval date equals its extraction date carries no real date — it is
    of unknown age, not newly approved — and would re-enter the "recent competitor
    approvals" signal on every run, burying the genuinely new registrations this
    report exists to surface.

    No extractor currently stamps dates this way: Guatemala briefly did, until its
    issue date turned out to be recoverable as expiration minus the five-year term.
    This stays as a guard so the pattern can never silently corrupt the recent signal.
    """
    appr = (row.get("approval_date") or "").strip()
    return bool(appr) and appr == (row.get("extraction_date") or "").strip()


def recent_approvals(rows, days: int = RECENT_DAYS) -> list[dict]:
    cutoff = date.today().toordinal() - days
    out = []
    for r in rows:
        if (r.get("record_type") or "") != "APPROVAL":
            continue
        if _has_placeholder_date(r):
            continue
        d = _parse(r.get("approval_date"))
        if d and d.toordinal() >= cutoff:
            out.append(r)
    out.sort(key=lambda r: _parse(r.get("approval_date")) or date.min, reverse=True)
    return out


# ── Text summary (email body) ─────────────────────────────────────────────────
def text_summary(rows: list[dict]) -> str:
    real, presence = _split(rows)
    approvals = sum(1 for r in real if r.get("record_type") == "APPROVAL")
    submissions = sum(1 for r in real if r.get("record_type") == "SUBMISSION")
    by_country = Counter(r.get("country_code") for r in real)
    recent = recent_approvals(real)

    lines = [
        f"RPI Competitive Intelligence — {date.today():%d/%m/%Y}",
        "",
        f"Total records: {len(real)}  ({approvals} approvals, {submissions} submissions)",
        f"Countries with data: {', '.join(f'{c}={n}' for c, n in sorted(by_country.items()))}",
        f"Competitor approvals in the last {RECENT_DAYS} days: {len(recent)}",
    ]
    if recent:
        lines.append("")
        lines.append(f"Most recent approvals:")
        for r in recent[:10]:
            lines.append(f"  • {r.get('approval_date')}  {r.get('country_code')}  "
                         f"{r.get('molecule_search_term')}  —  {r.get('product_name') or ''} "
                         f"({r.get('applicant') or ''})")
    if presence:
        terms = sorted({r.get("molecule_search_term") for r in presence})
        lines.append("")
        lines.append(f"No registrations found (checked): {', '.join(terms)}")
    return "\n".join(lines)


# ── xlsx workbook ─────────────────────────────────────────────────────────────
def build_report(rows: list[dict], output_path: str) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    real, presence = _split(rows)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    bold = Font(bold=True)

    def style_header(ws, row_idx, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")

    wb = Workbook()

    # ── Sheet 1: Resumen ──
    ws = wb.active
    ws.title = "Resumen"
    ws["A1"] = f"RPI Competitive Intelligence — {date.today():%d/%m/%Y}"
    ws["A1"].font = title_font
    approvals = sum(1 for r in real if r.get("record_type") == "APPROVAL")
    submissions = sum(1 for r in real if r.get("record_type") == "SUBMISSION")
    kpis = [
        ("Total records", len(real)),
        ("Approvals", approvals),
        ("Submissions", submissions),
        ("Distinct molecules", len({r.get("molecule_search_term") for r in real})),
        ("Countries with data", len({r.get("country_code") for r in real})),
        (f"Approvals (last {RECENT_DAYS}d)", len(recent_approvals(real))),
        ("Molecules checked, no data", len({r.get("molecule_search_term") for r in presence})),
    ]
    r0 = 3
    for i, (label, val) in enumerate(kpis):
        ws.cell(row=r0 + i, column=1, value=label).font = bold
        ws.cell(row=r0 + i, column=2, value=val)

    # By country table
    cstart = r0 + len(kpis) + 2
    ws.cell(row=cstart, column=1, value="By country").font = title_font
    ws.cell(row=cstart + 1, column=1, value="Country")
    ws.cell(row=cstart + 1, column=2, value="Approvals")
    ws.cell(row=cstart + 1, column=3, value="Submissions")
    style_header(ws, cstart + 1, 3)
    by_country = defaultdict(lambda: [0, 0])
    for r in real:
        idx = 0 if r.get("record_type") == "APPROVAL" else 1
        by_country[r.get("country_code")][idx] += 1
    for i, cc in enumerate(sorted(by_country)):
        ws.cell(row=cstart + 2 + i, column=1, value=cc)
        ws.cell(row=cstart + 2 + i, column=2, value=by_country[cc][0])
        ws.cell(row=cstart + 2 + i, column=3, value=by_country[cc][1])

    # Recent approvals (the headline signal) on the right
    recents = recent_approvals(real)
    rcol = 5
    ws.cell(row=cstart, column=rcol, value=f"Competitor approvals — last {RECENT_DAYS} days").font = title_font
    rec_headers = ["approval_date", "country_code", "molecule_search_term", "product_name", "applicant"]
    for j, h in enumerate(rec_headers):
        ws.cell(row=cstart + 1, column=rcol + j, value=h)
    style_header(ws, cstart + 1, rcol + len(rec_headers) - 1)
    for i, r in enumerate(recents[:40]):
        for j, h in enumerate(rec_headers):
            ws.cell(row=cstart + 2 + i, column=rcol + j, value=r.get(h))
    for col, w in {"A": 26, "B": 12, "C": 13, "E": 12, "F": 12, "G": 26, "H": 40, "I": 30}.items():
        ws.column_dimensions[col].width = w

    # ── Sheet 2: Matriz Molécula×País ──
    ws2 = wb.create_sheet("Matriz Molécula×País")
    countries = sorted({r.get("country_code") for r in real if r.get("country_code")})
    molecules = sorted({r.get("molecule_search_term") for r in real if r.get("molecule_search_term")})
    matrix = defaultdict(int)
    for r in real:
        matrix[(r.get("molecule_search_term"), r.get("country_code"))] += 1
    ws2.cell(row=1, column=1, value="Molecule")
    for j, cc in enumerate(countries):
        ws2.cell(row=1, column=2 + j, value=cc)
    ws2.cell(row=1, column=2 + len(countries), value="TOTAL")
    style_header(ws2, 1, 2 + len(countries))
    for i, mol in enumerate(molecules):
        ws2.cell(row=2 + i, column=1, value=mol).font = bold
        total = 0
        for j, cc in enumerate(countries):
            n = matrix.get((mol, cc), 0)
            total += n
            if n:
                ws2.cell(row=2 + i, column=2 + j, value=n)
        ws2.cell(row=2 + i, column=2 + len(countries), value=total)
    ws2.column_dimensions["A"].width = 30
    ws2.freeze_panes = "B2"

    # ── Sheet 3: Datos ──
    ws3 = wb.create_sheet("Datos")
    for j, field in enumerate(CANONICAL_FIELDS):
        ws3.cell(row=1, column=1 + j, value=field)
    style_header(ws3, 1, len(CANONICAL_FIELDS))
    for i, r in enumerate(rows):
        for j, field in enumerate(CANONICAL_FIELDS):
            ws3.cell(row=2 + i, column=1 + j, value=r.get(field))
    ws3.freeze_panes = "A2"
    last_col = get_column_letter(len(CANONICAL_FIELDS))
    ws3.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"
    widths = {"registration_number": 24, "product_name": 40, "api": 22,
              "applicant": 28, "molecule_search_term": 22, "source_url": 30}
    for j, field in enumerate(CANONICAL_FIELDS):
        ws3.column_dimensions[get_column_letter(1 + j)].width = widths.get(field, 14)

    wb.save(output_path)
    return output_path


def generate(output_dir: str, rows: list[dict] | None = None) -> str | None:
    """Generate the xlsx report; return its path, or None if there is no data."""
    rows = rows if rows is not None else load_consolidated(output_dir)
    if not rows:
        return None
    path = os.path.join(output_dir, f"RPI_REPORT_{date.today():%d%m%Y}.xlsx")
    return build_report(rows, path)
